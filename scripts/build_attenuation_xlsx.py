#!/usr/bin/env python3
"""Build Attenuation_Calibration_v3_5.xlsx per methodology §6.

Six sheets:
  1. Summary
  2. Within-Force
  3. Cross-Force
  4. Empirical Computation
  5. Mechanism Adjustments
  6. Trend Census

Color conventions (xlsx skill):
  - Blue  : hardcoded inputs / scenario-adjustable
  - Black : formulas / derived
  - Green : cross-sheet links
  - Yellow: key assumption that needs attention
"""
import sys, json, itertools, math
from pathlib import Path

# M16 (July 2026 review): repo-relative — the old hardcoded sandbox path
# made this provenance script unrunnable anywhere else.
REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO))
from pulse.seed_trends import get_report_trends
from pulse.config import FORCES, CATEGORIES

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CALIB_JSON = REPO / 'data' / 'attenuation_calibration_v3_5.json'
OUT        = REPO / 'data' / 'Attenuation_Calibration_v3_5.xlsx'

calib = json.loads(CALIB_JSON.read_text())
trends = get_report_trends()

# ── Style helpers ──────────────────────────────────────────────────────
FONT_NAME = 'Arial'
BLUE   = Font(name=FONT_NAME, color='0000FF')
BLUE_B = Font(name=FONT_NAME, color='0000FF', bold=True)
BLACK  = Font(name=FONT_NAME, color='000000')
BLACK_B= Font(name=FONT_NAME, color='000000', bold=True)
GREEN  = Font(name=FONT_NAME, color='008000')
GREEN_B= Font(name=FONT_NAME, color='008000', bold=True)
WHITE_B= Font(name=FONT_NAME, color='FFFFFF', bold=True)
BASE   = Font(name=FONT_NAME)

HDR_FILL   = PatternFill('solid', start_color='1F4E78')   # dark blue band
SUB_FILL   = PatternFill('solid', start_color='DDEBF7')   # light blue
YEL_FILL   = PatternFill('solid', start_color='FFFF00')
GRAY_FILL  = PatternFill('solid', start_color='F2F2F2')
GREEN_FILL = PatternFill('solid', start_color='E2EFDA')

THIN = Side(border_style='thin', color='BFBFBF')
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal='center', vertical='center')
LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
RIGHT  = Alignment(horizontal='right',  vertical='center')

def hdr(cell, txt):
    cell.value = txt
    cell.font = WHITE_B
    cell.fill = HDR_FILL
    cell.alignment = CENTER
    cell.border = BORDER_ALL

def sub(cell, txt):
    cell.value = txt
    cell.font = BLACK_B
    cell.fill = SUB_FILL
    cell.alignment = CENTER
    cell.border = BORDER_ALL

def note(cell, txt):
    cell.value = txt
    cell.font = BASE
    cell.alignment = LEFT

wb = Workbook()

# ════════════════════════════════════════════════════════════════════════
# Sheet 1: Summary
# ════════════════════════════════════════════════════════════════════════
s = wb.active
s.title = 'Summary'

s['A1'] = 'PRISM v3.5 — Attenuation Calibration Summary'
s['A1'].font = Font(name=FONT_NAME, bold=True, size=14)
s.merge_cells('A1:F1')

s['A2'] = f'Calibration version: {calib["calibration_version"]}'
s['A2'].font = BASE
s.merge_cells('A2:F2')

# Metadata block
s['A4'] = 'Calibration metadata'
s['A4'].font = BLACK_B

meta = [
    ('Trend base',                    calib['n_trends'],                   'number'),
    ('Unique pairs evaluated',        calib['n_unique_pairs'],             'number'),
    ('J₀ baseline (random-pair mean)',calib['J0_baseline'],                'decimal4'),
    ('Trend-weighted mean attenuation',calib['trend_weighted_mean_attenuation'],'decimal4'),
    ('Within-force clamp',            '[0.10, 0.45]',                      'text'),
    ('Cross-force clamp',             '[0.00, 0.45]',                      'text'),
    ('Mechanism adjustment range',    '±0.03 … ±0.10',                     'text'),
]
for i, (label, val, kind) in enumerate(meta, start=5):
    s.cell(i, 1, label).font = BLACK
    c = s.cell(i, 2, val)
    c.font = BLUE
    if kind == 'decimal4': c.number_format = '0.0000'
    elif kind == 'number': c.number_format = '#,##0'

# Force count table
r = 14
s.cell(r, 1, 'Force distribution'); s.cell(r,1).font = BLACK_B
r += 1
hdr(s.cell(r,1),'Force'); hdr(s.cell(r,2),'n trends'); hdr(s.cell(r,3),'% of base')
r += 1
total = calib['n_trends']
fc_start = r
for f in FORCES:
    n = calib['force_counts'][f]
    s.cell(r,1,f).font = BLACK
    s.cell(r,1).border = BORDER_ALL
    c = s.cell(r,2,n); c.font = BLUE; c.number_format='#,##0'; c.alignment = RIGHT; c.border=BORDER_ALL
    pct = s.cell(r,3, f'=B{r}/B${fc_start + len(FORCES)}')
    pct.font = BLACK; pct.number_format='0.0%'; pct.alignment = RIGHT; pct.border=BORDER_ALL
    r += 1
# Total row
s.cell(r,1,'TOTAL').font = BLACK_B
s.cell(r,1).border=BORDER_ALL
tot = s.cell(r,2, f'=SUM(B{fc_start}:B{r-1})')
tot.font = BLACK_B; tot.number_format='#,##0'; tot.alignment = RIGHT; tot.border=BORDER_ALL
s.cell(r,3, f'=SUM(C{fc_start}:C{r-1})').font = BLACK_B
s.cell(r,3).number_format='0.0%'; s.cell(r,3).border=BORDER_ALL; s.cell(r,3).alignment = RIGHT

# Within-force final — comparison to v3.1
r += 3
s.cell(r,1,'Within-force overlap — v3.5 vs v3.1'); s.cell(r,1).font = BLACK_B
r += 1
for i,h in enumerate(['Force','v3.5 final','v3.1 final','Δ','Direction'], start=1):
    hdr(s.cell(r,i), h)
r += 1
v34_w = calib['within_force_final']
v31_w = calib['v3_1_within_force_final']
for f in FORCES:
    s.cell(r,1,f).font = BLACK; s.cell(r,1).border=BORDER_ALL
    a = s.cell(r,2, v34_w[f]); a.font=BLUE; a.number_format='0.000'; a.alignment=RIGHT; a.border=BORDER_ALL
    b = s.cell(r,3, v31_w[f]); b.font=BLUE; b.number_format='0.000'; b.alignment=RIGHT; b.border=BORDER_ALL
    d = s.cell(r,4, f'=B{r}-C{r}'); d.font=BLACK; d.number_format='+0.000;-0.000;0.000'; d.alignment=RIGHT; d.border=BORDER_ALL
    dir_ = s.cell(r,5, f'=IF(D{r}>0,"↑ higher",IF(D{r}<0,"↓ lower","flat"))'); dir_.font=BLACK; dir_.alignment=CENTER; dir_.border=BORDER_ALL
    r += 1

# Per-force effective attenuation comparison
r += 2
s.cell(r,1,'Per-force effective attenuation — v3.5 vs v3.1'); s.cell(r,1).font = BLACK_B
r += 1
for i,h in enumerate(['Force','Row-mean cross overlap','eff_att_NEW','v3.1 eff_att','Δ'], start=1):
    hdr(s.cell(r,i), h)
r += 1
v34 = calib['per_force_effective_attenuation']
v31 = calib['v3_1_per_force_effective_attenuation']
rm = calib['per_force_row_mean_cross_overlap']
eff_start = r
for f in FORCES:
    s.cell(r,1,f).font = BLACK; s.cell(r,1).border=BORDER_ALL
    a = s.cell(r,2, rm[f]); a.font=BLUE; a.number_format='0.0000'; a.alignment=RIGHT; a.border=BORDER_ALL
    # eff_att is formula 0.5 × (1 − row_mean)
    b = s.cell(r,3, f'=0.5*(1-B{r})'); b.font=BLACK; b.number_format='0.000'; b.alignment=RIGHT; b.border=BORDER_ALL
    c = s.cell(r,4, v31[f]); c.font=BLUE; c.number_format='0.000'; c.alignment=RIGHT; c.border=BORDER_ALL
    d = s.cell(r,5, f'=C{r}-D{r}'); d.font=BLACK; d.number_format='+0.000;-0.000;0.000'; d.alignment=RIGHT; d.border=BORDER_ALL
    r += 1

# Trend-weighted mean sanity
r += 1
s.cell(r,1,'Trend-weighted mean attenuation (sanity check)').font = BLACK_B
# row_means are needed to compute weighted mean via formula
# Use formula referencing the eff-att block above
wm_cell = s.cell(r,2,
    f'=SUMPRODUCT(B{fc_start}:B{fc_start+len(FORCES)-1},C{eff_start}:C{eff_start+len(FORCES)-1})/B{fc_start+len(FORCES)}')
wm_cell.font = BLACK_B; wm_cell.number_format = '0.0000'; wm_cell.alignment = RIGHT
note(s.cell(r,4), 'Expected: 0.4492 ± rounding vs v3.1 0.446')

# Column widths
for col, w in [(1,38),(2,16),(3,16),(4,12),(5,16),(6,40)]:
    s.column_dimensions[get_column_letter(col)].width = w

# ════════════════════════════════════════════════════════════════════════
# Sheet 2: Within-Force
# ════════════════════════════════════════════════════════════════════════
s2 = wb.create_sheet('Within-Force')
s2['A1'] = 'Within-Force Overlap — Empirical + Mechanism'
s2['A1'].font = Font(name=FONT_NAME, bold=True, size=14)
s2.merge_cells('A1:H1')

# Method box
s2['A3'] = 'Method: mean weighted Jaccard across all (t,t\') pairs within a force → excess-over-baseline → ± mechanism adj → clamp [0.10, 0.45]'
s2['A3'].font = BASE; s2['A3'].alignment = LEFT
s2.merge_cells('A3:H3')
s2.row_dimensions[3].height = 30

# Table
r = 5
for i,h in enumerate(['Force','n','raw mean J','excess (J − J₀)/(1−J₀)','mech adj','excess + mech','FINAL (clamped)','Rationale for mechanism adj'], start=1):
    hdr(s2.cell(r,i), h)
s2.column_dimensions['H'].width = 60
r += 1

wraw = calib['within_force_raw_J']
wex  = calib['within_force_excess']
wme  = calib['within_force_mechanism_adj']
wfi  = calib['within_force_final']
J0   = calib['J0_baseline']

rationale_within = {
    'Consumer':      'Consumer trends span premium / sustainability / demographics / occasions / geography / value-trading — little mechanism clustering. Expanded by ultra-fast-fashion; diversity preserved. Minor −0.03 anti-cluster.',
    'Customer':      'Retail-media, agentic-commerce, DTC-pivot, live-commerce — channels partially coexist but each is a distinct buying motion. No adjustment — empirical excess sufficient.',
    'Technology':    'AI cluster (8 trends) + bio/chem cluster (5) + neurocosmetics tie in bio/chem. Mild +0.05 to recognize AI-stack mechanism clustering beyond structural score.',
    'Government':    'European Green Deal as common root: PFAS, Microplastics, PPWR, EUDR, Green Claims, AI Act, MoCRA. Adds +0.05 to already-high empirical signal — strongest coupling by construction.',
    'Environmental': 'Climate + water + palm + packaging mass balance — share substrate economics. +0.05 for supply-chain mechanism coupling.',
    'Competitive':   'Each trend is about a *different* competitor (P&G, Unilever, L\'Oréal, Reckitt, K-beauty, Amazon, AfCFTA local champions). Genuine diversity. −0.05 reflects structural distinction.',
}
data_start = r
for f in FORCES:
    n = calib['force_counts'][f]
    s2.cell(r,1,f).font = BLACK; s2.cell(r,1).border=BORDER_ALL; s2.cell(r,1).alignment=LEFT
    c = s2.cell(r,2,n); c.font=BLUE; c.number_format='#,##0'; c.alignment=RIGHT; c.border=BORDER_ALL
    c = s2.cell(r,3,wraw[f]); c.font=BLUE; c.number_format='0.0000'; c.alignment=RIGHT; c.border=BORDER_ALL
    # excess as formula using J0 from Summary? Keep as hardcoded blue (matches JSON) + reference note
    c = s2.cell(r,4,wex[f]); c.font=BLUE; c.number_format='0.0000'; c.alignment=RIGHT; c.border=BORDER_ALL
    c = s2.cell(r,5,wme[f]); c.font=BLUE; c.number_format='+0.00;-0.00;0.00'; c.alignment=RIGHT; c.border=BORDER_ALL
    c = s2.cell(r,6, f'=D{r}+E{r}'); c.font=BLACK; c.number_format='0.000'; c.alignment=RIGHT; c.border=BORDER_ALL
    c = s2.cell(r,7, f'=MAX(0.10, MIN(0.45, F{r}))'); c.font=BLACK_B; c.number_format='0.000'; c.alignment=RIGHT; c.border=BORDER_ALL
    c = s2.cell(r,8, rationale_within[f]); c.font=BASE; c.alignment=LEFT; c.border=BORDER_ALL
    s2.row_dimensions[r].height = 48
    r += 1

# Notes
r += 2
s2.cell(r,1,'Notes').font = BLACK_B
r += 1
notes_text = [
    f'J₀ (random-pair baseline) = {J0:.4f}. Excess transform: (J−J₀)/(1−J₀), clamped ≥ 0.',
    'Floor 0.10 prevents structurally-orthogonal trends in same force from summing undampened (cumulative-mass sanity).',
    'Ceiling 0.45 prevents any single force from losing all contribution in cumulative-mass calc.',
    'Consumer and Competitive floor-bind; four empirical excess signals are > 0 and drive the final value.',
]
for t in notes_text:
    s2.cell(r,1,'• ').font = BLACK
    s2.cell(r,2,t).font = BASE; s2.cell(r,2).alignment = LEFT
    s2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    s2.row_dimensions[r].height = 18
    r += 1

for col, w in [(1,14),(2,8),(3,12),(4,22),(5,10),(6,16),(7,16),(8,60)]:
    s2.column_dimensions[get_column_letter(col)].width = w

# ════════════════════════════════════════════════════════════════════════
# Sheet 3: Cross-Force
# ════════════════════════════════════════════════════════════════════════
s3 = wb.create_sheet('Cross-Force')
s3['A1'] = 'Cross-Force Overlap Matrix (FINAL, v3.5)'
s3['A1'].font = Font(name=FONT_NAME, bold=True, size=14)
s3.merge_cells('A1:H1')
s3['A2'] = 'Read as: row-force signal covered by column-force. Asymmetric by construction.'
s3['A2'].font = BASE

# 6x6 matrix
r = 4
hdr(s3.cell(r,1), '↓ From  /  To →')
for j, f in enumerate(FORCES, start=2):
    hdr(s3.cell(r, j), f)
r += 1
cf = calib['cross_force_final']
for i, fi in enumerate(FORCES):
    hdr(s3.cell(r, 1), fi)
    for j, fj in enumerate(FORCES, start=2):
        v = cf[fi][fj] if fi != fj else None
        c = s3.cell(r, j, v if v is not None else '—')
        if v is None:
            c.fill = GRAY_FILL; c.font = BLACK; c.alignment = CENTER
        else:
            c.font = BLUE; c.number_format = '0.000'; c.alignment = RIGHT
        c.border = BORDER_ALL
    r += 1

# Top 12 decomposed
r += 2
s3.cell(r,1,'Top 12 cross-force couplings — full decomposition').font = BLACK_B
r += 1
for i,h in enumerate(['Rank','From','To','raw J','excess','× asymm (√n_j/n_i, cap 1.5)','mech adj','FINAL (clamped 0–0.45)','Rationale'], start=1):
    hdr(s3.cell(r,i), h)
r += 1
flat = []
for fi in FORCES:
    for fj in FORCES:
        if fi == fj: continue
        flat.append((fi, fj,
                     calib['cross_force_raw_J'][fi][fj],
                     calib['cross_force_excess'][fi][fj],
                     calib['cross_force_asymm'][fi][fj],
                     calib['cross_force_mechanism_adj'][fi][fj],
                     calib['cross_force_final'][fi][fj]))
flat.sort(key=lambda x: -x[6])

rationale_cross = {
    ('Environmental','Government'): 'PFAS, PPWR, EUDR, Green Claims — environmental issues become regulatory.',
    ('Government','Environmental'): 'Same coupling, reverse direction; regulatory action targets env impact.',
    ('Customer','Government'):      'Retailer compliance scope expanding (GPSR, DPP, EUDR); channel carries reg burden.',
    ('Government','Customer'):      'Regulation mandates retailer disclosure / product passports / take-back.',
    ('Government','Technology'):    'Regulation triggers reformulation R&D (AI Act, MoCRA, PFAS reformulation).',
    ('Technology','Government'):    'Technology in regulatory scope (AI Act, Digital Product Passport).',
    ('Customer','Technology'):      'Retail media + agentic commerce = customer-tech hybrid.',
    ('Technology','Customer'):      'AI personalization reshapes retailer assortment and channel economics.',
    ('Environmental','Technology'): 'Supply constraints drive bio-chem / fermentation / palm substitutes.',
    ('Technology','Environmental'): 'New bio-chem / formulation reduces env footprint.',
    ('Customer','Environmental'):   'Retailer private-label pushes low-footprint positioning (e.g., Aldi, Lidl).',
    ('Environmental','Customer'):   'Sustainability claims reshape retailer gate-keeping.',
}
for rank, (fi, fj, raw, ex, asym, mech, fin) in enumerate(flat[:12], 1):
    s3.cell(r,1,rank).font=BLACK; s3.cell(r,1).alignment=CENTER; s3.cell(r,1).border=BORDER_ALL
    s3.cell(r,2,fi).font=BLACK;  s3.cell(r,2).border=BORDER_ALL
    s3.cell(r,3,fj).font=BLACK;  s3.cell(r,3).border=BORDER_ALL
    c=s3.cell(r,4,raw);  c.font=BLUE; c.number_format='0.0000'; c.alignment=RIGHT; c.border=BORDER_ALL
    c=s3.cell(r,5,ex);   c.font=BLUE; c.number_format='0.0000'; c.alignment=RIGHT; c.border=BORDER_ALL
    c=s3.cell(r,6,asym); c.font=BLUE; c.number_format='0.0000'; c.alignment=RIGHT; c.border=BORDER_ALL
    c=s3.cell(r,7,mech); c.font=BLUE; c.number_format='+0.00;-0.00;0.00'; c.alignment=RIGHT; c.border=BORDER_ALL
    c=s3.cell(r,8, f'=MAX(0, MIN(0.45, F{r}+G{r}))'); c.font=BLACK_B; c.number_format='0.000'; c.alignment=RIGHT; c.border=BORDER_ALL
    rat = rationale_cross.get((fi,fj), '—')
    c=s3.cell(r,9, rat); c.font=BASE; c.alignment=LEFT; c.border=BORDER_ALL
    s3.row_dimensions[r].height = 36
    r += 1

for col, w in [(1,6),(2,14),(3,14),(4,10),(5,10),(6,28),(7,10),(8,14),(9,55)]:
    s3.column_dimensions[get_column_letter(col)].width = w

# ════════════════════════════════════════════════════════════════════════
# Sheet 4: Empirical Computation
# ════════════════════════════════════════════════════════════════════════
s4 = wb.create_sheet('Empirical Computation')
s4['A1'] = 'Empirical Computation — Formulas & Worked Example'
s4['A1'].font = Font(name=FONT_NAME, bold=True, size=14)
s4.merge_cells('A1:F1')

r = 3
s4.cell(r,1,'Formulas').font = BLACK_B
r += 1
formulas = [
    ('Weighted Jaccard',        'J(x,y) = Σ min(e_x,c, e_y,c) / Σ max(e_x,c, e_y,c)  for c in 12 FMCG categories'),
    ('Random-pair baseline',    'J₀ = mean J over all unordered pairs in the full 99-trend base'),
    ('Excess-over-baseline',    'e(J̄) = max(0, (J̄ − J₀) / (1 − J₀))'),
    ('Asymmetric normalization','asymm_ij = e_ij × min(1.5, √(n_j / n_i))       [cross-force only]'),
    ('Final (within-force)',    'O_ii_final = clamp[0.10, 0.45]( e_ii + mech_ii )'),
    ('Final (cross-force)',     'O_ij_final = clamp[0.00, 0.45]( asymm_ij + mech_ij )'),
    ('Effective attenuation',   'eff_att_i = 0.5 × (1 − mean_{j ≠ i} O_ij_final)'),
]
for label, fx in formulas:
    s4.cell(r,1,label).font = BLACK_B
    s4.cell(r,2, fx).font = BASE
    s4.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    s4.row_dimensions[r].height = 20
    r += 1

# Worked example: pick two trends with distinct exposures
r += 2
s4.cell(r,1,'Worked example — weighted Jaccard on two trends').font = BLACK_B
r += 1
# pick consumer_r01 and government_r01
ex_a = next(t for t in trends if t.id == 'consumer_r01')
ex_b = next(t for t in trends if t.id == 'government_r01')

s4.cell(r,1,'Trend A').font = BLACK_B
s4.cell(r,2, f'{ex_a.id} — {ex_a.name}').font=BASE; s4.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
r += 1
s4.cell(r,1,'Trend B').font = BLACK_B
s4.cell(r,2, f'{ex_b.id} — {ex_b.name}').font=BASE; s4.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
r += 2

for i,h in enumerate(['Category','e_A','e_B','min','max'], start=1):
    hdr(s4.cell(r,i), h)
r += 1
min_start = r
for cat in CATEGORIES:
    ea = ex_a.category_exposure.get(cat, 0)
    eb = ex_b.category_exposure.get(cat, 0)
    s4.cell(r,1,cat).font=BLACK; s4.cell(r,1).border=BORDER_ALL
    c=s4.cell(r,2,ea); c.font=BLUE; c.alignment=RIGHT; c.border=BORDER_ALL
    c=s4.cell(r,3,eb); c.font=BLUE; c.alignment=RIGHT; c.border=BORDER_ALL
    c=s4.cell(r,4, f'=MIN(B{r},C{r})'); c.font=BLACK; c.alignment=RIGHT; c.border=BORDER_ALL
    c=s4.cell(r,5, f'=MAX(B{r},C{r})'); c.font=BLACK; c.alignment=RIGHT; c.border=BORDER_ALL
    r += 1
s4.cell(r,1,'Σ').font = BLACK_B; s4.cell(r,1).border=BORDER_ALL
c=s4.cell(r,4, f'=SUM(D{min_start}:D{r-1})'); c.font=BLACK_B; c.alignment=RIGHT; c.border=BORDER_ALL
c=s4.cell(r,5, f'=SUM(E{min_start}:E{r-1})'); c.font=BLACK_B; c.alignment=RIGHT; c.border=BORDER_ALL
sum_row = r
r += 2
s4.cell(r,1,'J (this pair)').font = BLACK_B
c = s4.cell(r,2, f'=D{sum_row}/E{sum_row}'); c.font=BLACK_B; c.number_format='0.0000'; c.alignment=RIGHT
r += 1
s4.cell(r,1,'J₀ baseline').font = BLACK_B
c = s4.cell(r,2, J0); c.font = BLUE; c.number_format='0.0000'; c.alignment=RIGHT
j0_ref = r
r += 1
s4.cell(r,1,'Excess transform').font = BLACK_B
c = s4.cell(r,2, f'=MAX(0,(B{r-2}-B{j0_ref})/(1-B{j0_ref}))'); c.font=BLACK_B; c.number_format='0.0000'; c.alignment=RIGHT

for col, w in [(1,28),(2,12),(3,12),(4,10),(5,10),(6,40)]:
    s4.column_dimensions[get_column_letter(col)].width = w

# ════════════════════════════════════════════════════════════════════════
# Sheet 5: Mechanism Adjustments
# ════════════════════════════════════════════════════════════════════════
s5 = wb.create_sheet('Mechanism Adjustments')
s5['A1'] = 'Mechanism Adjustments — Cell-by-Cell Rationale'
s5['A1'].font = Font(name=FONT_NAME, bold=True, size=14)
s5.merge_cells('A1:D1')

s5['A3'] = 'Adjustment range ±0.03 … ±0.10. Additive; applied after empirical/asymm transform; final value clamped.'
s5['A3'].font = BASE
s5.merge_cells('A3:D3')

# Within-force
r = 5
s5.cell(r,1,'Within-force adjustments').font = BLACK_B
r += 1
for i,h in enumerate(['Force','Adj','Sign','Rationale'], start=1):
    hdr(s5.cell(r,i), h)
r += 1
for f in FORCES:
    s5.cell(r,1,f).font = BLACK; s5.cell(r,1).border=BORDER_ALL
    c = s5.cell(r,2, wme[f]); c.font=BLUE; c.number_format='+0.00;-0.00;0.00'; c.alignment=RIGHT; c.border=BORDER_ALL
    sign = 'cluster' if wme[f] > 0 else ('anti-cluster' if wme[f] < 0 else 'neutral')
    s5.cell(r,3, sign).font=BLACK; s5.cell(r,3).alignment=CENTER; s5.cell(r,3).border=BORDER_ALL
    s5.cell(r,4, rationale_within[f]).font=BASE; s5.cell(r,4).alignment=LEFT; s5.cell(r,4).border=BORDER_ALL
    s5.row_dimensions[r].height = 40
    r += 1

# Cross-force non-zero
r += 2
s5.cell(r,1,'Cross-force adjustments (non-zero only)').font = BLACK_B
r += 1
for i,h in enumerate(['From','To','Adj','Rationale'], start=1):
    hdr(s5.cell(r,i), h)
r += 1
cm = calib['cross_force_mechanism_adj']
for fi in FORCES:
    for fj in FORCES:
        if fi == fj: continue
        adj = cm[fi][fj]
        if abs(adj) < 1e-9: continue
        s5.cell(r,1,fi).font=BLACK; s5.cell(r,1).border=BORDER_ALL
        s5.cell(r,2,fj).font=BLACK; s5.cell(r,2).border=BORDER_ALL
        c=s5.cell(r,3,adj); c.font=BLUE; c.number_format='+0.00;-0.00;0.00'; c.alignment=RIGHT; c.border=BORDER_ALL
        s5.cell(r,4, rationale_cross.get((fi,fj),'—')).font=BASE; s5.cell(r,4).alignment=LEFT; s5.cell(r,4).border=BORDER_ALL
        s5.row_dimensions[r].height = 32
        r += 1

for col, w in [(1,16),(2,16),(3,10),(4,70)]:
    s5.column_dimensions[get_column_letter(col)].width = w

# ════════════════════════════════════════════════════════════════════════
# Sheet 6: Trend Census
# ════════════════════════════════════════════════════════════════════════
s6 = wb.create_sheet('Trend Census')
s6['A1'] = f'Trend Census — {calib["n_trends"]} trends × 12 FMCG categories (exposure scores 0–5)'
s6['A1'].font = Font(name=FONT_NAME, bold=True, size=14)
s6.merge_cells('A1:P1')

r = 3
headers = ['#','Trend ID','Force','Name'] + CATEGORIES + ['Σ exposure']
for i,h in enumerate(headers, start=1):
    hdr(s6.cell(r,i), h)
r += 1
data_start = r

# sort by force then id for readability
force_order = {f:i for i,f in enumerate(FORCES)}
trends_sorted = sorted(trends, key=lambda t: (force_order.get(t.force, 99), t.id))
for idx, t in enumerate(trends_sorted, start=1):
    s6.cell(r,1,idx).font = BLACK; s6.cell(r,1).alignment=CENTER; s6.cell(r,1).border=BORDER_ALL
    s6.cell(r,2,t.id).font = BLACK; s6.cell(r,2).border=BORDER_ALL
    s6.cell(r,3,t.force).font = BLACK; s6.cell(r,3).border=BORDER_ALL
    s6.cell(r,4,t.name).font = BASE; s6.cell(r,4).border=BORDER_ALL; s6.cell(r,4).alignment=LEFT
    for j, cat in enumerate(CATEGORIES):
        v = t.category_exposure.get(cat, 0)
        c = s6.cell(r, 5+j, v); c.font = BLUE; c.alignment=RIGHT; c.number_format='0'; c.border=BORDER_ALL
    # Σ via formula
    sum_col = 5 + len(CATEGORIES)
    c = s6.cell(r, sum_col, f'=SUM({get_column_letter(5)}{r}:{get_column_letter(5+len(CATEGORIES)-1)}{r})')
    c.font = BLACK; c.alignment=RIGHT; c.number_format='0'; c.border=BORDER_ALL
    r += 1

# Column widths
s6.column_dimensions['A'].width = 5
s6.column_dimensions['B'].width = 18
s6.column_dimensions['C'].width = 14
s6.column_dimensions['D'].width = 46
for j in range(len(CATEGORIES)):
    s6.column_dimensions[get_column_letter(5+j)].width = 10
s6.column_dimensions[get_column_letter(5+len(CATEGORIES))].width = 12

# Freeze panes
s6.freeze_panes = 'E4'
s.freeze_panes  = 'A3'
s2.freeze_panes = 'A6'
s3.freeze_panes = 'B5'

wb.save(OUT)
print(f'Wrote {OUT}')
