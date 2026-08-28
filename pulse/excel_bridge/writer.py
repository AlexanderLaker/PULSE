"""Shift Matrix Excel writer — exports PRISM results to Excel.

Writes continuous-path shift matrices with percentile distributions, the 3D
regional shift matrix (F1, 2.10.0) and velocity data. All values are
percentages — no financial data. (The allocation sheet was removed with the
optimizer, D4 June 2026; force_attribution.direct_effects was deleted, F9
2.10.0.)
"""

import logging
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from pulse.config import ModelConfig, FORCES

logger = logging.getLogger(__name__)

# Styling
HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
HEADER_FONT = Font(name="Inter", color="F8FAFC", bold=True, size=10)
DATA_FONT = Font(name="Inter", color="1E293B", size=10)
EXPANSION_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
CONTRACTION_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
NEUTRAL_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


class ShiftMatrixWriter:
    """Writes PRISM simulation results to Excel."""

    def __init__(self, config: ModelConfig):
        self.config = config

    def write(self, output_path: str, mc_result: dict,
              allocation: dict = None,  # retired (D4); accepted+ignored for caller compat
              metadata: dict = None):
        """
        Write complete PRISM output to Excel.

        Creates sheets:
        1. Shift Matrix — continuous paths with percentiles
        2. Velocity & Triggers — path dynamics
        3. Metadata — run configuration and diagnostics
        """
        wb = openpyxl.Workbook()

        # Sheet 1: Shift Matrix (category roll-up)
        self._write_shift_matrix(wb, mc_result)

        # Sheet 2: Regional Shift (F1, 2.10.0 — the 3D category × region matrix)
        self._write_regional(wb, mc_result)

        # Sheet 3: Velocity & Triggers
        self._write_velocity(wb, mc_result)

        # Sheet 4: Metadata
        self._write_metadata(wb, mc_result, metadata)

        # Remove default sheet if extra sheets were created
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]


        wb.save(output_path)
        logger.info(f"Shift Matrix written to {output_path}")

    def _write_shift_matrix(self, wb, mc_result):
        """Write main shift matrix with continuous paths."""
        ws = wb.active
        ws.title = "Shift Matrix"

        shift_matrix = mc_result.get("shift_matrix", {})
        percentiles = ["p10", "p25", "median", "p75", "p90"]

        # Header
        ws.merge_cells("A1:A2")
        ws["A1"] = "Category"
        ws["A1"].font = HEADER_FONT
        ws["A1"].fill = HEADER_FILL
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        col = 2
        for year in self.config.path_years:
            start_col = col
            for p in percentiles:
                cell = ws.cell(row=2, column=col, value=p)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center")
                col += 1
            # Merge year header
            ws.merge_cells(start_row=1, start_column=start_col,
                           end_row=1, end_column=col - 1)
            ws.cell(row=1, column=start_col, value=str(year))
            ws.cell(row=1, column=start_col).font = HEADER_FONT
            ws.cell(row=1, column=start_col).fill = HEADER_FILL
            ws.cell(row=1, column=start_col).alignment = Alignment(horizontal="center")

        # Data rows
        row = 3
        for cat in self.config.category_names:
            cat_data = shift_matrix.get(cat, {})
            path = cat_data.get("path", {})

            ws.cell(row=row, column=1, value=cat).font = Font(name="Inter", bold=True, size=10)

            col = 2
            for year in self.config.path_years:
                year_data = path.get(year, {})
                for p in percentiles:
                    val = year_data.get(p, 0.0)
                    cell = ws.cell(row=row, column=col, value=round(val, 6))
                    cell.number_format = "0.00%"
                    cell.font = DATA_FONT
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = BORDER

                    # Color coding
                    if val > 0.005:
                        cell.fill = EXPANSION_FILL
                    elif val < -0.005:
                        cell.fill = CONTRACTION_FILL
                    else:
                        cell.fill = NEUTRAL_FILL

                    col += 1
            row += 1

        # Auto-fit columns
        for c in range(1, col):
            ws.column_dimensions[get_column_letter(c)].width = 12
        ws.column_dimensions["A"].width = 18

    def _write_regional(self, wb, mc_result):
        """F1 (2.10.0): the 3D shift resolved by (category, region), median per
        year. Rows are category × region; a globally-present trend reproduces
        the category number, so a region's row shows where the shift concentrates.
        """
        from pulse.config import REGIONS
        ws = wb.create_sheet("Regional Shift")
        rsm = mc_result.get("regional_shift_matrix", {})
        weights = mc_result.get("region_weights_used", {})

        ws.cell(row=1, column=1, value="Category").font = HEADER_FONT
        ws.cell(row=1, column=1).fill = HEADER_FILL
        ws.cell(row=1, column=2, value="Region (GP1 wt)").font = HEADER_FONT
        ws.cell(row=1, column=2).fill = HEADER_FILL
        for j, year in enumerate(self.config.path_years, 3):
            c = ws.cell(row=1, column=j, value=str(year))
            c.font = HEADER_FONT; c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center")

        row = 2
        for cat in self.config.category_names:
            by_region = rsm.get(cat, {})
            for region in REGIONS:
                wt = weights.get(region)
                label = f"{region}" + (f" ({wt*100:.0f}%)" if isinstance(wt, (int, float)) else "")
                ws.cell(row=row, column=1, value=cat).font = Font(name="Inter", size=9)
                ws.cell(row=row, column=2, value=label).font = Font(name="Inter", size=9)
                path = (by_region.get(region, {}) or {}).get("path", {})
                for j, year in enumerate(self.config.path_years, 3):
                    yd = path.get(year, {}) or path.get(str(year), {})
                    val = yd.get("median", 0.0) if isinstance(yd, dict) else 0.0
                    cell = ws.cell(row=row, column=j, value=round(val, 6))
                    cell.number_format = "0.00%"
                    cell.font = DATA_FONT
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = BORDER
                    if val > 0.005:
                        cell.fill = EXPANSION_FILL
                    elif val < -0.005:
                        cell.fill = CONTRACTION_FILL
                    else:
                        cell.fill = NEUTRAL_FILL
                row += 1

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 20
        for c in range(3, 3 + len(self.config.path_years)):
            ws.column_dimensions[get_column_letter(c)].width = 11

    def _write_velocity(self, wb, mc_result):
        """Write velocity and trigger analysis."""
        ws = wb.create_sheet("Velocity & Triggers")
        shift_matrix = mc_result.get("shift_matrix", {})

        # Header
        headers = ["Category", "Path Shape"] + [f"Velocity {y}" for y in self.config.path_years[1:]]
        for j, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=j, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        from pulse.common.shape_compat import velocity_median as _vel_median

        row = 2
        for cat in self.config.category_names:
            cat_data = shift_matrix.get(cat, {})
            velocity = cat_data.get("velocity", {})

            ws.cell(row=row, column=1, value=cat).font = Font(name="Inter", bold=True)

            # Path shape classification (simplified)
            vel_vals = [_vel_median(v) for v in velocity.values()]
            if vel_vals:
                if all(abs(v) < 0.001 for v in vel_vals):
                    shape = "flat"
                elif abs(vel_vals[-1]) > abs(vel_vals[0]) * 1.5:
                    shape = "accelerating"
                elif abs(vel_vals[0]) > abs(vel_vals[-1]) * 1.5:
                    shape = "decelerating"
                else:
                    shape = "gradual"
            else:
                shape = "n/a"
            ws.cell(row=row, column=2, value=shape)

            for j, year in enumerate(self.config.path_years[1:], 3):
                val = _vel_median(velocity.get(year, 0.0))
                cell = ws.cell(row=row, column=j, value=round(val, 6))
                cell.number_format = "0.00%"
            row += 1

    def _write_metadata(self, wb, mc_result, extra_metadata=None):
        """Write run metadata and configuration."""
        ws = wb.create_sheet("Metadata")

        # v3.2: per-force attenuation. No flat 0.5 default exists anywhere.
        # The Config sheet (built separately) carries the full overlap
        # detail; here we surface each force's effective value plus the
        # unweighted mean for at-a-glance context.
        pfa = dict(getattr(self.config, "per_force_attenuation", {}))
        force_order = ["Consumer", "Customer", "Technology", "Government",
                       "Environmental", "Competitive"]
        pfa_str = ", ".join(
            f"{f}={pfa[f]:.3f}" for f in force_order if f in pfa
        ) or "n/a"
        pfa_mean = (
            sum(pfa.values()) / len(pfa) if pfa else 0.0
        )

        # D17 (owner decision, June 2026): the attenuation provenance is
        # labeled "structured-judgment overlap correction", never
        # "calibrated" — the values rest on an exposure-overlap proxy plus
        # documented judgment adjustments, not measured outcomes (F-19).
        att_source = (
            "admin override"
            if getattr(self.config, "attenuation_source", "") == "admin_override"
            else "structured-judgment overlap correction (v3.5, Apr-2026)"
        )

        data = [
            ("PRISM Shift Matrix", ""),
            ("Generated", datetime.now().isoformat()),
            # D21 sweep: version comes from the engine result — the previous
            # hardcoded "2.5.0" string had drifted three releases behind.
            ("Model Version", str(mc_result.get("model_version", "unknown"))
                              + " — Bayesian MC, Gaussian copula"),
            ("Numerics Backend", str(mc_result.get("numerics_backend", "n/a"))),  # D13
            ("Iterations", mc_result.get("iterations", "N/A")),
            ("Model Type", mc_result.get("model_type", "N/A")),
            ("Attenuation Source", att_source),
            ("Per-Force Attenuation", pfa_str),
            ("Per-Force Mean (unweighted)", f"{pfa_mean:.3f}"),
            ("Path Years", str(self.config.path_years)),
            ("", ""),
            # D16 (owner decision, June 2026): design assumption travels with
            # every export of the matrix.
            ("READING NOTE", "Ceteris paribus: assumes no management response — no pricing"),
            ("", "moves, innovation, or competitive reaction by Henkel or competitors."),
            ("", "Totals read as exposure if nobody acts, not as forecast outcomes."),
            ("", ""),
            # 2.9.0: VC-lens basis travels with every export of the run.
            ("VC ATTRIBUTION", "Epicentre partition: each trend's contribution is assigned"),
            ("", "wholly to the value-chain stage where experts located its impact"),
            ("", "epicentre. Propagation up/down the chain is not modelled."),
            ("", ""),
            # F1 (2.10.0): the structural scale chain + the regional roll-up.
            ("SCALE (F1)", "Structural pass-through per force = force weight (1/6) x attenuation"),
            ("", "(~0.40-0.50) ~= 7%. A near-certain trend touching 20% of a category's"),
            ("", "GP1 moves it ~1.4% at full materialization, not 20% — read cells as a"),
            ("", "conservative, comparable index and apply with GP1_proj = GP1 x (1+shift)."),
            ("REGIONAL (F1)", "The shift math is 3D (category x region x year): a trend hits a"),
            ("", "(category, region) cell weighted by category exposure x regional exposure."),
            ("", "Category numbers are the region-GP1-weighted roll-up (see Region weights);"),
            ("", "a regionally-concentrated trend only moves its regions' slice of the pool."),
            ("REGION WEIGHTS", ", ".join(
                f"{r} {w*100:.0f}%" for r, w in (mc_result.get("region_weights_used") or {}).items()
            ) or "equal (fallback)"),
            ("", "Proxy: Henkel Group FY2025 regional sales split (HCB not disclosed by region)."),
            ("", ""),
            # F4/F5 (2.10.0): timing uncertainty + the probability-score meaning.
            ("TIMING (F4)", "P10-P90 bands are magnitude uncertainty (fixed Beta concentration"),
            ("", "a+b=6). Per-iteration peak-year jitter (+/-1yr triangular) gives the"),
            ("", "velocity bands their timing content. start_year gates the onset (F11)."),
            ("SCORE->PRIOR (F5)", "1->0.17  2->0.33  3->0.50  4->0.67  5->0.83 (Beta mean = score/6;"),
            ("", "deliberate shrinkage vs overconfidence — a '5' is five-in-six, not certainty)."),
            # F8 (2.10.0): copula scale.
            ("COPULA (F8)", "Configured correlations are LATENT-scale (co-movement of the underlying"),
            ("", "drivers); realized score-score correlation is slightly lower (~0.28 at 0.30)."),
            ("", ""),
            ("SECURITY NOTE", "This file contains ONLY percentage shifts."),
            ("", "No company financial data (NES, GP1, GP2) is present."),
            ("", "Apply shifts to your financials in your own Excel."),
        ]

        if extra_metadata:
            data.append(("", ""))
            for k, v in extra_metadata.items():
                data.append((str(k), str(v)))

        for row_idx, (key, val) in enumerate(data, 1):
            ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True) if key else DATA_FONT
            ws.cell(row=row_idx, column=2, value=val)

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 60
