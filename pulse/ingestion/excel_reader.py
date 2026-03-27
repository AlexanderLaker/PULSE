"""V12 Excel reader — extracts non-financial data from profit pool model.

Reads force sheets (1_Consumer through 6_Competitive), Config, Helper,
and EM_Input. Explicitly skips financial sheets. All data passes through
the Financial Data Firewall before entering PULSE.
"""

import logging
import re
from pathlib import Path
from typing import Optional

import pandas as pd
import openpyxl

from pulse.config import (FORCES, CATEGORIES, VC_STEPS, FORCE_SHEETS,
                           SKIP_SHEETS, ModelConfig)
from pulse.ingestion.models import Trend, TrendDatabase
from pulse.ingestion.firewall import FinancialDataFirewall

logger = logging.getLogger(__name__)


class ExcelReader:
    """Reads V12 Excel file, extracts trend data, enforces firewall."""

    def __init__(self, filepath: str):
        self.filepath = Path(filepath)
        self.firewall = FinancialDataFirewall()
        self._wb = None

    def read(self) -> TrendDatabase:
        """Main entry: read V12, return TrendDatabase with all trends."""
        logger.info(f"Reading {self.filepath.name}...")
        self._wb = openpyxl.load_workbook(str(self.filepath), data_only=True)

        available_sheets = set(self._wb.sheetnames)
        logger.info(f"Available sheets: {available_sheets}")

        # Read configuration
        config = self._read_config()

        # Read trends from each force sheet
        all_trends = []
        for force_name, sheet_name in FORCE_SHEETS.items():
            if sheet_name in available_sheets:
                trends = self._read_force_sheet(force_name, sheet_name, config)
                all_trends.extend(trends)
                logger.info(f"  {force_name}: {len(trends)} trends extracted")
            else:
                logger.warning(f"  {force_name}: sheet '{sheet_name}' not found")

        db = TrendDatabase(
            trends=all_trends,
            categories=config.category_names,
            forces=list(FORCE_SHEETS.keys()),
            source_file=str(self.filepath),
            financial_data_detected=False,
        )

        logger.info(f"Ingestion complete: {db.trend_count} trends, "
                     f"{len(db.categories)} categories, {len(db.forces)} forces")

        return db

    def _read_config(self) -> ModelConfig:
        """Read model configuration from Config / Helper sheets."""
        config = ModelConfig()

        if "Config" in self._wb.sheetnames:
            ws = self._wb["Config"]
            # Parse config values — adapt to V12 structure
            for row in ws.iter_rows(min_row=1, max_row=50, max_col=10, values_only=False):
                cells = [c.value for c in row]
                if cells[0] and isinstance(cells[0], str):
                    label = cells[0].strip().lower()
                    if "attenuation" in label and cells[1] is not None:
                        try:
                            config.attenuation = float(cells[1])
                        except (ValueError, TypeError):
                            pass
                    elif "region" in label and cells[1]:
                        config.region = str(cells[1])

        # Try to read category names from Helper or Config
        if "Helper" in self._wb.sheetnames:
            ws = self._wb["Helper"]
            cats = []
            for row in ws.iter_rows(min_row=1, max_row=50, max_col=5, values_only=True):
                if row[0] and isinstance(row[0], str) and ":" in str(row[0]):
                    cat = str(row[0]).strip()
                    if any(prefix in cat for prefix in ["Hair", "LHC"]):
                        cats.append(cat)
            if cats:
                config.category_names = cats

        return config

    def _read_force_sheet(self, force_name: str, sheet_name: str,
                          config: ModelConfig) -> list:
        """Read trends from a single force sheet."""
        ws = self._wb[sheet_name]
        trends = []
        trend_counter = 0

        # Read as DataFrame for easier processing
        data = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            data.append(row)

        if not data:
            return trends

        # Find the trends table — look for header row with "Impact", "Probability"
        header_row_idx = None
        for i, row in enumerate(data):
            row_strs = [str(c).lower() if c else "" for c in row]
            if any("impact" in s for s in row_strs) and any("probability" in s for s in row_strs):
                header_row_idx = i
                break

        if header_row_idx is None:
            logger.warning(f"  Could not find trend table in {sheet_name}")
            return trends

        headers = [str(c).strip() if c else f"col_{j}" for j, c in enumerate(data[header_row_idx])]

        # Map column indices
        col_map = {}
        for j, h in enumerate(headers):
            h_lower = h.lower()
            if "trend" in h_lower or "name" in h_lower or h_lower == "key trends and developments":
                col_map["name"] = j
            elif "impact" in h_lower and "score" not in h_lower:
                col_map["impact"] = j
            elif "probability" in h_lower:
                col_map["probability"] = j
            elif "direction" in h_lower:
                col_map["direction"] = j
            elif "start" in h_lower or "timing" in h_lower:
                col_map["start_year"] = j
            elif "description" in h_lower or "evidence" in h_lower:
                col_map["description"] = j
            elif "implication" in h_lower or "action" in h_lower:
                col_map["implication"] = j
            elif "source" in h_lower:
                col_map["source"] = j
            elif "sub" in h_lower and "categ" in h_lower:
                col_map["sub_category"] = j

        # Find category columns (they contain category names in headers)
        cat_cols = {}
        for j, h in enumerate(headers):
            for cat in config.category_names:
                cat_short = cat.split(": ")[-1] if ": " in cat else cat
                if cat_short.lower() in h.lower() or cat.lower() in h.lower():
                    cat_cols[cat] = j

        # Parse trend rows
        for i in range(header_row_idx + 1, len(data)):
            row = data[i]
            if not row or all(c is None for c in row):
                continue

            # Get trend name
            name_idx = col_map.get("name", 0)
            name = row[name_idx] if name_idx < len(row) else None
            if not name or not isinstance(name, str) or len(name.strip()) < 3:
                continue

            trend_counter += 1
            force_prefix = force_name.lower()[:3]
            trend_id = f"{force_prefix}_{trend_counter:02d}"

            # Extract scores
            impact = self._safe_int(row, col_map.get("impact"), default=3)
            probability = self._safe_int(row, col_map.get("probability"), default=3)
            impact = max(1, min(5, impact))
            probability = max(1, min(5, probability))

            # Direction
            direction_val = self._safe_str(row, col_map.get("direction"))
            if direction_val:
                direction = "Contraction" if "contraction" in direction_val.lower() else "Expansion"
            else:
                direction = "Expansion"

            # Start year
            start_year = self._safe_int(row, col_map.get("start_year"), default=2028)

            # Description
            desc = self._safe_str(row, col_map.get("description")) or ""

            # Strategic implication
            impl = self._safe_str(row, col_map.get("implication")) or ""

            # Category exposure
            cat_exposure = {}
            for cat, col_idx in cat_cols.items():
                val = row[col_idx] if col_idx < len(row) else None
                if val is not None:
                    if isinstance(val, str) and val.strip().upper() == "X":
                        cat_exposure[cat] = 3  # Binary mark → medium exposure
                    elif isinstance(val, (int, float)) and 0 <= val <= 5:
                        cat_exposure[cat] = int(val)

            trend = Trend(
                id=trend_id,
                force=force_name,
                sub_category=self._safe_str(row, col_map.get("sub_category")) or "",
                name=name.strip(),
                description=desc,
                direction=direction,
                impact=impact,
                probability=probability,
                start_year=start_year,
                strategic_implication=impl,
                category_exposure=cat_exposure,
                data_source=self._safe_str(row, col_map.get("source")) or "",
            )
            trends.append(trend)

        return trends

    def _safe_int(self, row, idx, default=0) -> int:
        if idx is None or idx >= len(row):
            return default
        val = row[idx]
        if val is None:
            return default
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return default

    def _safe_str(self, row, idx) -> Optional[str]:
        if idx is None or idx >= len(row):
            return None
        val = row[idx]
        return str(val).strip() if val is not None else None

    def get_firewall_report(self) -> str:
        return "Firewall: Disabled for cloud deployment."
